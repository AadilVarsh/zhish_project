import openpyxl
from openpyxl import utils
import os

#E:\Varsh_Code\Python\MEET_BOT\venv\Scripts\python.exe E:/Varsh_Code/Python/zhish_project/excel_handler/pyxlhndlr.py
#/e/Varsh_Code/Python/zhish_project/data_handler

def __init__():
    pass

def sheet(wb_path):
        wb = openpyxl.load_workbook(wb_path)
        sheet = wb.sheetnames
        sheet = wb.active

        return sheet

def get_sheet_format(sheet):
    if sheet.max_column == 2:
        return "custom"
    elif sheet.max_column == 3:
        return "monthly"

def get_msg_info(sheet_format, sheet):
    user_info = []
    user_custom = {
            "name" : str(),
            "phone_number" : "+" + str()
            }

    user_monthly = {
            "name" : str(),
            "phone_number" : "+" + str(),
            "last_service" : str()
            }

    if sheet_format == "custom":
        for row in range(sheet.max_row):
            user_custom["name"] , user_custom["phone_number"] = sheet.cell(row = row + 1, column = 1).value, sheet.cell(row = row + 1, column = 2).value

            user_info.append([user_custom["name"], user_custom['phone_number']])
        

    if sheet_format == "monthly":
        for row in range(sheet.max_row):
            user_monthly['name'], user_monthly['phone_number'], user_monthly['last_service'] = sheet.cell(row = row + 1, column = 1).value, sheet.cell(row =row + 1, column = 2).value, sheet.cell(row = row + 1, column = 3).value

            user_info.append([user_monthly["name"], user_monthly["phone_number"], user_monthly["last_service"]])
    return user_info

        

if __name__ == "__main__":
    sheet = sheet(r"E:\Varsh_Code\Python\zhish_project\workbooks\testXL.xlsx")
    print(sheet.max_row)
    print(get_sheet_format(sheet = sheet))
    print(get_msg_info(get_sheet_format(sheet= sheet), sheet))

    
